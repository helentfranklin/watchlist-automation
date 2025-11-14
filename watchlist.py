import yfinance as yf
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from datetime import datetime

WATCHLIST = ["AAPL", "MSFT", "NVDA", "AMZN", "META"]
OUTPUT_FILE = "daily_watchlist.xlsx"

GREEN = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
RED = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

def fetch_data(ticker):
    try:
        data = yf.Ticker(ticker).history(period="1d")
        if data.empty:
            return None
        latest = data.iloc[-1]
        return {
            "Ticker": ticker,
            "Price": float(latest["Close"]),
            "Open": float(latest["Open"]),
            "High": float(latest["High"]),
            "Low": float(latest["Low"]),
            "Volume": int(latest["Volume"]),
        }
    except:
        return None

def generate_excel():
    wb = Workbook()
    ws = wb.active
    ws.title = "Watchlist"
    headers = ["Ticker", "Price", "Open", "High", "Low", "Volume"]
    ws.append(headers)

    for ticker in WATCHLIST:
        row_data = fetch_data(ticker)
        if row_data is None:
            ws.append([ticker, "N/A", "N/A", "N/A", "N/A", "N/A"])
            continue

        row = list(row_data.values())
        ws.append(row)

        price = row_data["Price"]
        open_price = row_data["Open"]
        excel_row = ws.max_row

        if price > open_price:
            for col in range(1, 7):
                ws.cell(excel_row, col).fill = GREEN
        elif price < open_price:
            for col in range(1, 7):
                ws.cell(excel_row, col).fill = RED

    wb.save(OUTPUT_FILE)

if __name__ == "__main__":
    generate_excel()
    print("Watchlist Excel generated at", datetime.now())
